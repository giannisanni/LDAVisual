from openpyxl import load_workbook

# Paths to your Excel files
source_file_path = 'C:/Users/gsanr/PycharmProjects/LDA_topic_modeling/export (1) completely sorted - year wise updated-2 - Copy2.xlsx'
destination_file_path = 'C:/Users/gsanr/PycharmProjects/LDA_topic_modeling/export (1) completely sorted - year wise updated-2 - Copy.xlsx'

# Load the source and destination workbooks
source_workbook = load_workbook(filename=source_file_path)
destination_workbook = load_workbook(filename=destination_file_path)

# Copy each sheet from the source workbook to the destination workbook
for sheet_name in source_workbook.sheetnames:
    # Get the sheet from the source workbook
    source_sheet = source_workbook[sheet_name]

    # Create a new sheet in the destination workbook with the same name
    if sheet_name not in destination_workbook.sheetnames:
        destination_sheet = destination_workbook.create_sheet(sheet_name)
    else:
        destination_sheet = destination_workbook[sheet_name]

    # Copy data from each cell of the source sheet to the destination sheet
    for row in source_sheet.iter_rows():
        for cell in row:
            destination_sheet[cell.coordinate].value = cell.value

# Save the destination workbook
destination_workbook.save(filename=destination_file_path)

# Close both workbooks
source_workbook.close()
destination_workbook.close()
