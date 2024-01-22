from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import requests
import pdfplumber
import os

# List of DG names to search for
dg_names = [
    "Luchtvaart en Maritieme",
    "Mobiliteit",
    "Rijkswaterstaat",
    "Water en Bodem",
    "Milieu en Internationaal"
]

# Function to extract hyperlink from a cell
def extract_hyperlink(cell):
    if cell.hyperlink:
        return cell.hyperlink.target
    else:
        return None

# Function to extract DG name from the top right of the first page of a PDF
def extract_dg_name(file_path, dg_names):
    with pdfplumber.open(file_path) as pdf:
        first_page = pdf.pages[0]
        top_right = (
            first_page.width / 2, 0, first_page.width, first_page.height / 4
        )
        text = first_page.crop(top_right).extract_text()
        for dg_name in dg_names:
            if dg_name in text:
                return dg_name
        return "Not Found"

# Load the workbook
excel_file_path = 'C:/Users/gsanr/PycharmProjects/LDA_topic_modeling/export (1) completely sorted - year wise updated-2 - Copy4.xlsx'
workbook = load_workbook(filename=excel_file_path)

# Set up the WebDriver
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

# Process each sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    dg_column_index = sheet.max_column + 1

    for row_number, cell in enumerate(sheet['A'], start=1):  # Assuming 'Titel' is in the first column
        if cell.hyperlink:
            url = cell.hyperlink.target
            try:
                driver.get(url)
                pdf_link_element = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "h2.result--title.icon.icon--large.icon--publication > a"))
                )
                pdf_link = pdf_link_element.get_attribute('href')

                response = requests.get(pdf_link)
                temp_pdf_path = 'C:/Users/gsanr/PycharmProjects/LDA_topic_modeling/temp_pdf.pdf'
                with open(temp_pdf_path, 'wb') as f:
                    f.write(response.content)

                dg_name = extract_dg_name(temp_pdf_path, dg_names)
                sheet.cell(row=row_number, column=dg_column_index, value=dg_name)

                os.remove(temp_pdf_path)

            except (NoSuchElementException, TimeoutException):
                print(f"PDF link not found for URL {url}")
                sheet.cell(row=row_number, column=dg_column_index, value="PDF Link Not Found")
            except Exception as e:
                print(f"Error processing URL {url}: {e}")
                sheet.cell(row=row_number, column=dg_column_index, value="Error")

            time.sleep(0.5)

# Save the workbook with the new columns
workbook.save(filename=excel_file_path)

# Close the WebDriver and the workbook
driver.quit()
workbook.close()
